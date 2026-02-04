"""
Excel Import/Export Utilities
Functions to handle Excel file import and export
"""

import pandas as pd
import io
from typing import List, Dict, Any, Tuple
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.schemas import TransactionBase


def parse_excel_date(date_value: Any) -> datetime.date:
    """
    Parse various date formats from Excel

    Args:
        date_value: Date value from Excel (can be string, datetime, etc.)

    Returns:
        Parsed date object
    """
    if pd.isna(date_value):
        return None

    if isinstance(date_value, datetime):
        return date_value.date()

    if isinstance(date_value, str):
        # Try DD.MM.YY format first
        for fmt in ["%d.%m.%y", "%d.%m.%Y", "%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"]:
            try:
                return datetime.strptime(date_value.strip(), fmt).date()
            except ValueError:
                continue

    return None


def import_excel_file(file_content: bytes, sheet_name: str = None) -> Dict[str, Any]:
    """
    Import transactions from Excel file

    Args:
        file_content: Excel file content as bytes
        sheet_name: Optional sheet name to import (default: first sheet)

    Returns:
        Dictionary with account_name, opening_balance, and transactions list
    """
    # Read Excel file
    excel_file = pd.ExcelFile(io.BytesIO(file_content))

    # Use specified sheet or first sheet
    if sheet_name and sheet_name in excel_file.sheet_names:
        sheet = sheet_name
    else:
        sheet = excel_file.sheet_names[0]

    df = pd.read_excel(excel_file, sheet_name=sheet, header=None)

    # Extract account name from first row
    account_name = str(df.iloc[0, 0]).replace(" TRACKER", "").strip()

    # Extract opening balance from row 2 (index 2)
    opening_balance = Decimal(str(df.iloc[2, 1])) if not pd.isna(df.iloc[2, 1]) else Decimal("0")

    # Find the header row (contains "Date", "Amount", etc.)
    header_row_idx = None
    for idx, row in df.iterrows():
        if idx < 3:
            continue
        if "Date" in str(row.iloc[0]) or "date" in str(row.iloc[0]).lower():
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise ValueError("Could not find header row in Excel file")

    # Read data starting from header row
    df_data = pd.read_excel(excel_file, sheet_name=sheet, header=header_row_idx)

    # Clean column names
    df_data.columns = df_data.columns.str.strip()

    # Find relevant columns (case-insensitive)
    date_col = None
    amount_col = None
    person_col = None
    narration_col = None

    for col in df_data.columns:
        col_lower = str(col).lower()
        if "date" in col_lower and date_col is None:
            date_col = col
        elif "amount" in col_lower and amount_col is None:
            amount_col = col
        elif "paid" in col_lower or "person" in col_lower or "from" in col_lower:
            person_col = col
        elif "narration" in col_lower or "description" in col_lower:
            narration_col = col

    if not all([date_col, amount_col, person_col]):
        raise ValueError("Required columns not found. Expected: Date, Amount, Paid To/From")

    # Parse transactions
    transactions = []

    for _, row in df_data.iterrows():
        # Skip rows with empty date or amount
        if pd.isna(row[date_col]) or pd.isna(row[amount_col]):
            continue

        # Skip rows where person is NaN
        if pd.isna(row[person_col]):
            continue

        try:
            trans_date = parse_excel_date(row[date_col])
            if trans_date is None:
                continue

            amount = Decimal(str(row[amount_col]))
            person = str(row[person_col]).strip()
            narration = str(row[narration_col]).strip() if narration_col and not pd.isna(row[narration_col]) else None

            transactions.append({
                "date": trans_date,
                "amount": amount,
                "paid_to_from": person,
                "narration": narration
            })
        except (ValueError, TypeError):
            # Skip invalid rows
            continue

    return {
        "account_name": account_name,
        "opening_balance": opening_balance,
        "transactions": transactions
    }


def export_to_excel(
    accounts_data: List[Dict[str, Any]],
    output_stream: io.BytesIO
) -> None:
    """
    Export accounts and transactions to Excel file

    Args:
        accounts_data: List of dictionaries containing account and transaction data
        output_stream: BytesIO stream to write Excel file to
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for account_data in accounts_data:
        account_name = account_data["name"]
        opening_balance = account_data["opening_balance"]
        transactions = account_data["transactions"]
        person_summary = account_data.get("person_summary", [])

        # Create worksheet
        ws = wb.create_sheet(title=account_name[:31])  # Excel sheet name limit

        # Header styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Row 1: Title
        ws["A1"] = f"{account_name.upper()} TRACKER"
        ws["A1"].font = Font(bold=True, size=14)
        ws["B1"] = opening_balance

        # Row 3: Opening Balance
        ws["A3"] = "Opening Balance:"
        ws["B3"] = opening_balance

        # Row 5: Column Headers
        headers = ["Date", "Amount (INR)", "Paid To/From", "Narration", "Balance (INR)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Transactions data
        current_row = 6
        for trans in transactions:
            ws.cell(row=current_row, column=1, value=trans["date"].strftime("%d.%m.%y"))
            ws.cell(row=current_row, column=2, value=float(trans["amount"]))
            ws.cell(row=current_row, column=3, value=trans["paid_to_from"])
            ws.cell(row=current_row, column=4, value=trans["narration"] or "")
            ws.cell(row=current_row, column=5, value=float(trans["balance_after"]))
            current_row += 1

        # Person summary sidebar (starting from column G)
        if person_summary:
            ws.cell(row=5, column=7, value="AUTO-TRACKING (150 NAMES)").font = header_font

            ws.cell(row=7, column=7, value="Person").font = header_font
            ws.cell(row=7, column=8, value="Total (INR)").font = header_font
            ws.cell(row=7, column=9, value="Count").font = header_font

            summary_row = 8
            for person in person_summary:
                ws.cell(row=summary_row, column=7, value=person["person"])
                ws.cell(row=summary_row, column=8, value=float(person["total_amount"]))
                ws.cell(row=summary_row, column=9, value=person["transaction_count"])
                summary_row += 1

        # Summary section at bottom
        summary_start_row = max(current_row + 2, 20)

        ws.cell(row=summary_start_row, column=7, value="SUMMARY").font = Font(bold=True)
        ws.cell(row=summary_start_row + 1, column=7, value="Current:")
        ws.cell(row=summary_start_row + 1, column=8, value=float(account_data.get("current_balance", 0)))

        ws.cell(row=summary_start_row + 2, column=7, value="Total Out:")
        ws.cell(row=summary_start_row + 2, column=8, value=float(account_data.get("total_expense", 0)))

        ws.cell(row=summary_start_row + 3, column=7, value="Total In:")
        ws.cell(row=summary_start_row + 3, column=8, value=float(account_data.get("total_income", 0)))

        ws.cell(row=summary_start_row + 4, column=7, value="Unique Names:")
        ws.cell(row=summary_start_row + 4, column=8, value=account_data.get("unique_persons", 0))

        ws.cell(row=summary_start_row + 6, column=7, value="Capacity:")
        ws.cell(row=summary_start_row + 6, column=8, value="150 names")

        ws.cell(row=summary_start_row + 7, column=7, value="Transactions:")
        ws.cell(row=summary_start_row + 7, column=8, value=f"{len(transactions)} rows")

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["G"].width = 25
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 10

    # Save to stream
    wb.save(output_stream)
    output_stream.seek(0)
